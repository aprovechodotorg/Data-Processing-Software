#v0.4  Python3, 8 september 2023
    #v0.4: added load_timeseries_with_uncertainty() and write_timeseries_with_uncertainty()

#    Copyright (C) 2023 Aprovecho Research Center 
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.    
#
#    Contact: sam@aprovecho.org

from uncertainties import ufloat
from uncertainties import unumpy
import csv
from openpyxl import load_workbook
import xlrd
import math
import os
import time
from datetime import datetime as dt
import numpy as np
import re

#This is a library of functions for LEMS-Data-Processing for input and output files. The input functions read input files and store the data in dictionaries. The output functions copy the data dictionaries to an output file. 

#####################################################################
def load_inputs_from_spreadsheet(Inputpath):
    #if cell value is blank don't read it in
    #do: add case for opening xls files using xlrd
    
    #function reads in spreadsheet (data entry form) and stores variable names, units, and values in dictionaries
    #Input: Inputpath: spreadsheet file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_TE_DataEntryForm.xlsx
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    val={}      #dictionary keys are variable names, values are variable values
    unc={}  #dictionary keys are variable names, values are variable uncertainty values
    uval={}  #dictionary keys are variable names, values are variable ufloats (val,unc pair)
    
    #make header line and store in dictionary
    name='variable_name'
    names.append(name)
    units[name] = 'units'
    val[name]= 'value'
    unc[name]='uncertainty'
    
    wb = load_workbook(filename = Inputpath, data_only=True)    #load spreadsheet
    sheet=wb.active #grab first sheet
    
    #iterate through all cells in the sheet. Find 'label' as reference point to read in cells
    grabvals = 0    #flag to read in cells after 'label' is found
    colnum=0    #initialize column number
    for col in sheet.iter_cols():   #for each column in the sheet
        colnum=colnum+1
        rownum=0    #initialize row number
        for cell in col:    #for each cell in the column
            rownum = rownum+1
            if grabvals == 1:   #if the cell should be read in
                if cell.value is None:  #if cell is blank then stop reading in cell values
                    grabvals = 0
                else:   #if cell is not blank then read it in
                    name=cell.value
                    names.append(name)
                    units[name] = sheet.cell(row=rownum, column=units_colnum).value 
                    val[name] = sheet.cell(row=rownum, column=colnum-1).value    #variable value is one cell to the left of the label
                    #val[name] = sheet.cell(row=rownum, column=colnum-2).value    #if spreadsheet includes uncertainty cells, variable value is 2 cells left of label
                    #unc[name] = sheet.cell(row=rownum, column=colnum-1).value     #if spreadsheet includes uncertainty cells, uncertainty value is 1 cell left of label            
            if cell.value == 'label':
                grabvals = 1 #start reading in cells
                #find the units column (the location varies)
                for n in range(colnum,0,-1):    #for each column to the left of label
                    nextcell=sheet.cell(row=rownum, column=n).value #read the cell
                    if nextcell in ['Units','units']:   #if it is the units column
                        units_colnum= n                 #record the column number
                        break
                        
    return names,units,val,unc              #type: list, dict, dict, dict
#####################################################################
def load_constant_inputs(Inputpath):
    #function loads in variables from csv input file and stores variable names, units, and values in dictionaries
    #Input: Inputpath: csv file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_EnergyInputs.csv
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    val={}      #dictionary keys are variable names, values are variable values
    unc={}  #dictionary keys are variable names, values are variable uncertainty values
    uval={}  #dictionary keys are variable names, values are ufloat variables (val,unc pair) 
    
    #load input file
    stuff=[]
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    #put inputs in a dictionary    
    for row in stuff:
        name=row[0]
        units[name]=row[1]
        val[name]=row[2]
        try:                                #some input files may not have row 3
            unc[name]=row[3] 
        except:
            unc[name]=''
        try:
            float(val[name])           #if val is a number
            try:
                float(unc[name])   #if unc is a number
                uval[name]=ufloat(float(val[name]),float(unc[name]))
            except:
                uval[name]=ufloat(row[2],0)
        except:     #if val is not a number, but rather a string
            uval[name]=row[2] #uval is a string
        names.append(name)
            
    return names,units,val,unc,uval
#######################################################################
def load_timeseries_with_header(Inputpath):
    #function loads in raw time series data csv input file from sensor box with header and startup diagnostics. Stores variable names, units, header parameters, and time series data in dictionaries
    #Input: Inputpath: csv file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawData.csv
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    A={}      #dictionary keys are variable names, values are A parameters (span)
    B={}  #dictionary keys are variable names, values are B parameters (offset)
    C={}  #dictionary keys are variable names, values are C parameters (constant variable names)
    D={}  #dictionary keys are variable names, values are D parameters (constant variable values)
    const = {} #dictionary keys are constant variable names(C parameters), values are constant variable values (D parameters)
    data = {} #dictionary keys are variable names, values are time series as a list
    
    #load input file
    stuff=[]
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    #find the row indices for the header and the data
    for n,row in enumerate(stuff[:100]): #iterate through first 101 rows to look for the header
        if row[0] == '#A:':
            Arow = n
        if row[0] == '#B:':
            Brow = n 
        if row[0] == '#C:':
            Crow = n
        if row[0] == '#D:':
            Drow = n      
        if row[0] == '#units:':
            unitsrow = n
        if row[0] == 'time':
            namesrow = n

    datarow = namesrow + 1
        
    names=stuff[namesrow]    
    for n,name in enumerate(names):
        units[name]=stuff[unitsrow][n]
        data[name]=[x[n] for x in stuff[datarow:]]
        for m,val in enumerate(data[name]):
            try: 
                data[name][m]=float(data[name][m])
            except:
                pass
        try:
            A[name]=float(stuff[Arow][n])
        except:  
            A[name]=stuff[Arow][n]        
        try:
            B[name]=float(stuff[Brow][n])
        except:
            B[name]=stuff[Brow][n]   
        try:
            C[name]=float(stuff[Crow][n])
        except:
            C[name]=stuff[Crow][n]
        try:
            D[name]=float(stuff[Drow][n])
        except:
            D[name]=stuff[Drow][n]

        #define the constant parameters (names are C parameters, values are D parameters)
        if type(C[name]) is str:
            const[C[name]] = D[name]   
         
    return names,units,data,A,B,C,D,const

#######################################################################

def load_header(Inputpath):
    #function loads in header from raw time series data csv input file or header input file. Stores variable names, units, header parameters in dictionaries
    # same as load_timeseries_with_header() but without data series
    #Input: Inputpath: csv file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawData.csv
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    A={}      #dictionary keys are variable names, values are A parameters (span)
    B={}  #dictionary keys are variable names, values are B parameters (offset)
    C={}  #dictionary keys are variable names, values are C parameters (constant variable names)
    D={}  #dictionary keys are variable names, values are D parameters (constant variable values)
    const = {} #dictionary keys are constant variable names(C parameters), values are constant variable values (D parameters)
    
    #load input file
    stuff=[]
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    #find the row indices for the header and the data
    for n,row in enumerate(stuff[:100]): #iterate through first 101 rows to look for the header
        if row[0] == '#A:':
            Arow = n
        if row[0] == '#B:':
            Brow = n 
        if row[0] == '#C:':
            Crow = n
        if row[0] == '#D:':
            Drow = n      
        if row[0] == '#units:':
            unitsrow = n
        if row[0] == 'time':
            namesrow = n
        
    names=stuff[namesrow]    
    for n,name in enumerate(names):
        units[name]=stuff[unitsrow][n]
        try:
            A[name]=float(stuff[Arow][n])
        except:  
            A[name]=stuff[Arow][n]        
        try:
            B[name]=float(stuff[Brow][n])
        except:
            B[name]=stuff[Brow][n]   
        try:
            C[name]=float(stuff[Crow][n])
        except:
            C[name]=stuff[Crow][n]
        try:
            D[name]=float(stuff[Drow][n])
        except:
            D[name]=stuff[Drow][n]

        #define the constant parameters (names are C parameters, values are D parameters)
        if type(C[name]) is str:
            const[C[name]] = D[name]   
         
    return names,units,A,B,C,D,const
    
##########################################################################################

def load_timeseries(Inputpath):
    #function loads in time series data from csv input file and stores variable names, units, and time series in dictionaries
    #Input: Inputpath: csv file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawData.csv
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    data = {} #dictionary keys are variable names, values are time series as a list
    
    #load input file
    stuff=[]
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    names=stuff[0]    #first row is channel names
    for n,name in enumerate(names):
        units[name]=stuff[1][n] #second row is units
        data[name]=[x[n] for x in stuff[2:]]    #data series
        for m,val in enumerate(data[name]):
            try: 
                data[name][m]=float(data[name][m])
            except:
                pass

    return names,units,data
#######################################################################
def load_timeseries_with_uncertainty(Inputpath):
    #function loads in time series data from csv input file and stores variable names, units, and time series in dictionaries
    #Input: Inputpath: csv file to load. example: C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawData.csv
    
    names = [] #list of variable names
    units={}    #dictionary keys are variable names, values are units
    data = {} #dictionary keys are variable names, values are time series as a list
    nom={}
    unc={}
    
    #load input file
    stuff=[]
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    ucnames=stuff[0]    #first row is channel names
    for n,name in enumerate(ucnames[:-1]):
        if '_uc' not in name:
            units[name]=stuff[1][n] #second row is units
            names.append(name)
            nom[name]=[x[n] for x in stuff[2:]]    #data series
            unc[name]=[x[n+1] for x in stuff[2:]]    #data series
            data[name]=[]
        
            for m,val in enumerate(nom[name]):
                try:
                    float(val)           #if val is a number
                    try:
                        float(unc[name][m])   #if unc is a number
                        uval=ufloat(float(val),float(unc[name][m]))
                    except:
                        uval=ufloat(float(val),0)
                except:     #if val is not a number, but rather a string
                    uval=val
                data[name].append(uval)

    return names,units,data
#################################################################

def write_constant_outputs(Outputpath,Names,Units,Val,Unc,Uval):
    #function writes output variables from dictionaries to csv output file
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_EnergyOutputs.csv
        #Names: list of variable names
        #units: dictionary keys are variable names, values are units
        #val: dictionary keys are variable names, values are variable values
        #nom: dictionary keys are variable names, values are variable nominal values 
        #unc: dictionary keys are variable names, values are variable uncertainty values
    
    #store data as a list of lists to print by row
    for name in Names:
        try:                                                    #see if a nominal value exists
            Val[name]
        except:                                             #if not then 
            try:                                                #try getting the nominal value from the ufloat
                Val[name]=Uval[name].n
            except:                                        #and if that doesn't work then define the nominal value as the single value
                Val[name]=Uval[name]
        try:                                                   #see if uncertainty value exists
            Unc[name]
        except:                                             #if not then
            try:                                                #try getting the uncertainty value from the ufloat
                Unc[name]=Uval[name].s
            except:
                Unc[name]=''                            #and if that doesn't work then define the uncertainty value as blank

    output=[]                                               #initialize list of lines
    for name in Names:                           #for each variable
        row=[]                                               #initialize row
        row.append(name)
        row.append(Units[name])
        row.append(Val[name])
        row.append(Unc[name])
        output.append(row)                          #add the row to output list

    #print to the output file
    with open(Outputpath,'w',newline='') as csvfile: 
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################

def write_timeseries_with_header(Outputpath,Names,Units,Data,A,B,C,D):
    #function writes time series data csv output file including raw data header with calibration parameters. All variables are taken from dictionaries. 
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
        #Names: list of variable names
        #Units: dictionary keys are channel names, values are units
        #Data: dictionary keys are channel names, values are time series as a list
        #A: dictionary keys are channel names, values are A parameters in header
        #B: dictionary keys are channel names, values are B parameters in header
        #C: dictionary keys are channel names, values are C parameters in header
        #D: dictionary keys are channel names, values are D parameters in header        
    
    #make lists for each header line
    Arow=[]
    Brow=[]
    Crow=[]
    Drow=[]
    Unitsrow=[] #initialize empty rows for the header
    for name in Names:
        Arow.append(A[name])
        Brow.append(B[name])
        Crow.append(C[name])
        Drow.append(D[name])
        Unitsrow.append(Units[name])
        
    #store data as a list of lists to print by row
    output=[Arow,Brow,Crow,Drow,Unitsrow,Names]          #initialize list of output lines starting with header
    for n,val in enumerate(Data['time']):   #for each data point in the time series
        row=[]                                                  #initialize blank row
        for name in Names:                          #for each channel
            row.append(Data[name][n])          #add the data point
        output.append(row)                              #add the row to the output list            
    
    #print to the output file
    with open(Outputpath,'w',newline='') as csvfile: 
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################

def write_header(Outputpath,Names,Units,A,B,C,D):
    #function writes raw data header to csv file. Same as write_timeseries_with_header() but without data series. All variables are taken from dictionaries. 
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
        #Names: list of variable names
        #Units: dictionary keys are channel names, values are units
        #A: dictionary keys are channel names, values are A parameters in header
        #B: dictionary keys are channel names, values are B parameters in header
        #C: dictionary keys are channel names, values are C parameters in header
        #D: dictionary keys are channel names, values are D parameters in header        
    
    #make lists for each header line
    Arow=[]
    Brow=[]
    Crow=[]
    Drow=[]
    Unitsrow=[] #initialize empty rows for the header
    for name in Names:
        Arow.append(A[name])
        Brow.append(B[name])
        Crow.append(C[name])
        Drow.append(D[name])
        Unitsrow.append(Units[name])
        
    #store data as a list of lists to print by row
    output=[Arow,Brow,Crow,Drow,Unitsrow,Names]          #initialize list of output lines starting with header
    
    #print to the output file
    with open(Outputpath,'w',newline='') as csvfile: 
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################

def write_timeseries(Outputpath,Names,Units,Data):
    #function writes time series data csv output file. All variables are taken from dictionaries. 
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
        #Names: list of variable names
        #Units: dictionary keys are channel names, values are units
        #Data: dictionary keys are channel names, values are time series as a list
    
    #make list for units row

    Unitsrow=[] #initialize empty row
    for name in Names:
        Unitsrow.append(Units[name])
        
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('unitsrow '+timestampstring)  
        
    #store data as a list of lists to print by row
    output=[Names,Unitsrow]          #initialize list of output lines starting with header
    try:
        for n,val in enumerate(Data['time']):   #for each data point in the time series
            row=[]                                                  #initialize blank row
            for name in Names:                          #for each channel
                row.append(Data[name][n])          #add the data point
            output.append(row)                              #add the row to the output list
    except:
        for n, val in enumerate(Data['time_test']):  # for each data point in the time series
            row = []  # initialize blank row
            for name in Names:  # for each channel
                row.append(Data[name][n])  # add the data point
            output.append(row)  # add the row to the output list
            
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('created output '+timestampstring)
    
    #print to the output file
    with open(Outputpath,'w',newline='') as csvfile: 
        writer = csv.writer(csvfile)
        for row in output:
            if n%100 == 0:
                timestampobject=dt.now()    #get timestamp from operating system
                timestampstring=timestampobject.strftime("%H:%M:%S")  
                print('n '+timestampstring)
            writer.writerow(row)
            
###################################################
           
def write_timeseries_with_uncertainty(Outputpath,Names,Units,Data):
    #function writes time series data csv output file. All variables are taken from dictionaries. 
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
        #Names: list of variable names
        #Units: dictionary keys are channel names, values are units
        #Data: dictionary keys are channel names, values are time series as a list
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('start write_timeseries_with_uncertainty '+timestampstring)
    newnames = []
    regex1 = re.compile('_smooth')
    regex2 = re.compile('_Ave')
    for name in Names:
        if re.search(regex1,name) or re.search(regex2,name):
            newnames.append(name)
            newnames.append(name+'_uc')
            Units[name+'_uc'] = Units[name]
            try:    #ufloats
                Data[name+'_uc'] = ['']*len(Data[name])
                Data[name] = unumpy.nominal_values(Data[name])
            except: #not ufloats
                Data[name+'_uc'] = ['']*len(Data[name])
            print("UC not printed ",name,", too slow")
        else:
            newnames.append(name)
            newnames.append(name+'_uc')
            Units[name+'_uc'] = Units[name]
            try:    #ufloats
                Data[name+'_uc'] = unumpy.std_devs(Data[name])
                Data[name] = unumpy.nominal_values(Data[name])
            except: #not ufloats
                Data[name+'_uc'] = ['']*len(Data[name])
            print(name)
            
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('updated Data '+timestampstring)
    
    write_timeseries(Outputpath,newnames,Units,Data)
    
    '''
    #check for the file
    if os.path.isfile(Outputpath):
        os.remove(Outputpath) #and remove it (because writing appends)
        
    #make list for names row and units row
    Namesrow = []   #initialize empty row
    Unitsrow=[] #initialize empty row
    for name in Names:
        Namesrow.append(name)
        Namesrow.append(name+'_uc')
        Unitsrow.append(Units[name])
        Unitsrow.append(Units[name]) 

    output=[Namesrow,Unitsrow]      #store data as a list of lists to print by row
    #print to the output file
    with open(Outputpath,'a',newline='') as csvfile: 
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
    output = []
    
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('saved header '+timestampstring)
    
    data_array = []
    for name in Names:
        try:    #ufloats
            data_array.append(unumpy.nominal_values(Data[name]))
            data_array.append(unumpy.std_devs(Data[name]))
        except: #not ufloats
            data_array.append(Data[name])
            data_array.append(['']*len(Data[name]))
  
    data_array = [list(sublist) for sublist in list(zip(*data_array))]  #transpose
    
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('created array '+timestampstring)  
    
    for n,array_row in enumerate(data_array):
        output.append(array_row)
        if n % 1000 == 0 or n == len(data_array)-1:
            timestampobject=dt.now()    #get timestamp from operating system
            timestampstring=timestampobject.strftime("%H:%M:%S")  
            print('row '+str(n)+' '+timestampstring)
            print(array_row)
    
    exit()
    
            #print to the output file
            with open(Outputpath,'a',newline='') as csvfile: 
                writer = csv.writer(csvfile)
                for row in output:
                    writer.writerow(row)
            output = []
    
            timestampobject=dt.now()    #get timestamp from operating system
            timestampstring=timestampobject.strftime("%H:%M:%S")  
            print('saved array '+timestampstring)
    
    data_array = np.zeros((len(Names)*2,len(Data[name])))
    for i,name in enumerate(Names):
        try:    #ufloats
            data_array[i] = np.array([unumpy.nominal_values(Data[name])])
            data_array[i+1] = np.array([unumpy.std_devs(Data[name])])
        except: #not ufloats
            data_array[i] = np.array([Data[name]])
            data_array[i+1] = np.array([['']*len(Data[name])])
            data_array = np.transpose(data_array)      
    
          
            

    
    #store data as a list of lists to print by row
    output=[Namesrow,Unitsrow]          #initialize list of output lines starting with header
    for n in range(len(Data['time'])):   #for each data point in the time series
        row=[]                                                  #initialize blank row
        for name in Names:                          #for each channel
            try:    #ufloat
                row.append(Data[name][n].n)          #add the data point
                row.append(Data[name][n].s)          #add the data point
            except: #not ufloat
                row.append(Data[name][n])          #add the data point
                row.append('')                                  #add the data point
        output.append(row)                              #add the row to the output list            
        if n % 1000 == 0 or n == len(Data['time'])-1:
            timestampobject=dt.now()    #get timestamp from operating system
            timestampstring=timestampobject.strftime("%H:%M:%S")  
            print('row '+str(n)+' '+timestampstring)
            
            #print to the output file
            with open(Outputpath,'a',newline='') as csvfile: 
                writer = csv.writer(csvfile)
                for outrow in output:
                    writer.writerow(outrow)
            output = []
            timestampobject=dt.now()    #get timestamp from operating system
            timestampstring=timestampobject.strftime("%H:%M:%S")  
            print('saved '+timestampstring)
            
            
            
    ##print to the output file
    #with open(Outputpath,'w',newline='') as csvfile: 
    #    writer = csv.writer(csvfile)
    #    for row in output:
    #        writer.writerow(row)
   ''' 
########################################################################
           
def write_timeseries_without_uncertainty(Outputpath,Names,Units,Data):
    # similar to write_timeseries_with_uncertainty but only takes the nominal value
    #function writes time series data csv output file. All variables are taken from dictionaries. 
    #Inputs:
        #Outputpath: output csv file that will be created. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
        #Names: list of variable names
        #Units: dictionary keys are channel names, values are units
        #Data: dictionary keys are channel names, values are time series as a list
    
    timestampobject=dt.now()    #get timestamp from operating system
    timestampstring=timestampobject.strftime("%H:%M:%S")  
    print('start write_timeseries_without_uncertainty '+timestampstring)
    
    #check for the file
    if os.path.isfile(Outputpath):
        os.remove(Outputpath) #and remove it (because writing appends)
        
    #make list for names row and units row
    Namesrow = []   #initialize empty row
    Unitsrow=[] #initialize empty row
    for name in Names:
        Namesrow.append(name)
        #Namesrow.append(name+'_uc')
        Unitsrow.append(Units[name])
        #Unitsrow.append(Units[name])    
    #store data as a list of lists to print by row
    output=[Namesrow,Unitsrow]          #initialize list of output lines starting with header
    for n in range(len(Data['time'])):   #for each data point in the time series
        row=[]                                                  #initialize blank row
        for name in Names:                          #for each channel
            try:    #ufloat
                row.append(Data[name][n].n)          #add the data point
                #row.append(Data[name][n].s)          #add the data point
            except: #not ufloat
                row.append(Data[name][n])          #add the data point
                #row.append('')                                  #add the data point
        output.append(row)                              #add the row to the output list            
        if n % 1000 == 0 or n == len(Data['time'])-1:
            #print to the output file
            with open(Outputpath,'a',newline='') as csvfile: 
                writer = csv.writer(csvfile)
                for outrow in output:
                    writer.writerow(outrow)
            output = []
            
            timestampobject=dt.now()    #get timestamp from operating system
            timestampstring=timestampobject.strftime("%H:%M:%S")  
            print(str(n)+' '+timestampstring)
            
    ##print to the output file
    #with open(Outputpath,'w',newline='') as csvfile: 
    #    writer = csv.writer(csvfile)
    #    for row in output:
    #        writer.writerow(row)
    
########################################################################
def write_logfile(Logpath,Logs):
    #writes to logfile.txt to document data manipulations
    #Inputs: 
    #Logpath: logfile.txt path. example:  C:\Mountain Air\equipment\Ratnoze\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_log.txt
    #Logs: list of lines that will get logged to the file
    with open(Logpath, 'a') as logfile: 
        for log in Logs:
            logfile.write('\n'+log)
#######################################################################
