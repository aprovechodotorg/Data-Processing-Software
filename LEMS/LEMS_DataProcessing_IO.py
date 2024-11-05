# v0.2  Python3

#    Copyright (C) 2022 Aprovecho Research Center 
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.    
#
#    Contact: sam@aprovecho.org

from uncertainties import ufloat
import csv
from openpyxl import load_workbook
from datetime import datetime as dt
from uncertainties import unumpy
import os
import re
import easygui
import math
import subprocess
from LEMS_RedoFirmwareCalcs import RedoFirmwareCalcs

# This is a library of functions for LEMS-Data-Processing for input and output files. The input functions read input
# files and store the data in dictionaries. The output functions copy the data dictionaries to an output file.

#####################################################################


def load_inputs_from_spreadsheet(Inputpath):
    # if cell value is blank don't read it in
    # do: add case for opening xls files using xlrd
    
    # function reads in spreadsheet (data entry form) and stores variable names, units, and values in dictionaries
    # Input: Inputpath: spreadsheet file to load. example:
    # Data/alcohol/alcohol_test1/alcohol_test1_TE_DataEntryForm.xlsx
    
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    val = {}   # dictionary keys are variable names, values are variable values
    unc = {}  # dictionary keys are variable names, values are variable uncertainty values
    uval = {}  # dictionary keys are variable names, values are variable ufloats (val,unc pair)
    
    # make header line and store in dictionary
    name = 'variable_name'
    names.append(name)
    units[name] = 'units'
    val[name] = 'value'
    unc[name] = 'uncertainty'
    
    wb = load_workbook(filename=Inputpath, data_only=True)  # load spreadsheet
    sheet = wb.active  # grab first sheet
    
    # iterate through all cells in the sheet. Find 'label' as reference point to read in cells
    grabvals = 0  # flag to read in cells after 'label' is found
    colnum = 0  # initialize column number
    for col in sheet.iter_cols():  # for each column in the sheet
        colnum = colnum+1
        rownum = 0   # initialize row number
        for cell in col:   # for each cell in the column
            rownum = rownum+1
            if grabvals == 1:   # if the cell should be read in
                if cell.value is None:  # if cell is blank then stop reading in cell values
                    grabvals = 0
                else:   # if cell is not blank then read it in
                    name = cell.value
                    names.append(name)
                    units[name] = sheet.cell(row=rownum, column=units_colnum).value 
                    val[name] = sheet.cell(row=rownum, column=colnum-1).value   # variable value is one cell to the
                    # left of the label
                    # val[name] = sheet.cell(row=rownum, column=colnum-2).value
                    # if spreadsheet includes uncertainty cells, variable value is 2 cells left of label
                    # unc[name] = sheet.cell(row=rownum, column=colnum-1).value
                    # if spreadsheet includes uncertainty cells, uncertainty value is 1 cell left of label
            if cell.value == 'label':
                grabvals = 1  # start reading in cells
                # find the units column (the location varies)
                for n in range(colnum, 0, -1):    # for each column to the left of label
                    nextcell = sheet.cell(row=rownum, column=n).value  # read the cell
                    if nextcell in ['Units', 'units']:   # if it is the units column
                        units_colnum = n   # record the column number
                        break
                        
    return names, units, val, unc              # type: # list, dict, dict, dict
#####################################################################


def load_constant_inputs(Inputpath):
    # function loads in variables from csv input file and stores variable names, units, and values in dictionaries
    # Input: Inputpath: csv file to load. example: Data/alcohol/alcohol_test1/alcohol_test1_EnergyInputs.csv
    
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    val = {}   # dictionary keys are variable names, values are variable values
    unc = {}  # dictionary keys are variable names, values are variable uncertainty values
    uval = {}  # dictionary keys are variable names, values are ufloat variables (val,unc pair)
    
    # load input file
    stuff = []
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    # put inputs in a dictionary
    for row in stuff:
        name = row[0]
        units[name] = row[1]
        val[name] = row[2]
        try:   # some input files may not have row 3
            unc[name] = row[3]
        except IndexError:
            unc[name] = ''
        try:
            float(val[name])   # if val is a number
            try:
                float(unc[name])   # if unc is a number
                uval[name] = ufloat(float(val[name]), float(unc[name]))
            except ValueError:
                uval[name] = ufloat(row[2], 0)
        except TypeError:  # if val is not a number, but rather a string
            uval[name] = row[2]  # uval is a string
        names.append(name)
            
    return names, units, val, unc, uval
#######################################################################


def load_timeseries_with_header(Inputpath):
    # function loads in raw time series data csv input file from sensor box with header and startup diagnostics.
    # Stores variable names, units, header parameters, and time series data in dictionaries
    # Input: Inputpath: csv file to load. example: Data/alcohol/alcohol_test1/alcohol_test1_RawData.csv
    
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    A = {}  # dictionary keys are variable names, values are A parameters (span)
    B = {}  # dictionary keys are variable names, values are B parameters (offset)
    C = {}  # dictionary keys are variable names, values are C parameters (constant variable names)
    D = {}  # dictionary keys are variable names, values are D parameters (constant variable values)
    const = {}  # dictionary keys are constant variable names(C parameters), values are constant variable values
    # (D parameters)
    data = {}  # dictionary keys are variable names, values are time series as a list
    
    # load input file
    stuff = []
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)

    version = ''

    # find the row indices for the header and the data
    for n, row in enumerate(stuff[:100]):  # iterate through first 101 rows to look for the header
        if row[0] == '#A:':
            Arow = n
        if row[0] == '#B:':
            Brow = n 
        if row[0] == '#C:':
            Crow = n
        if row[0] == '#D:':
            Drow = n      
        if row[0] == '#units:':
            unitsrow = n
        if row[0] == 'time':
            namesrow = n
        if '#version' in row[0]:
            version = row[0]

    datarow = namesrow + 1
        
    names = stuff[namesrow]
    for n, name in enumerate(names):
        units[name] = stuff[unitsrow][n]
        data[name] = [x[n] for x in stuff[datarow:]]
        for m, val in enumerate(data[name]):
            try: 
                data[name][m] = float(data[name][m])
            except ValueError:
                pass
        try:
            A[name] = float(stuff[Arow][n])
        except ValueError:
            A[name] = stuff[Arow][n]
        try:
            B[name] = float(stuff[Brow][n])
        except ValueError:
            B[name] = stuff[Brow][n]
        try:
            C[name] = float(stuff[Crow][n])
        except ValueError:
            C[name] = stuff[Crow][n]
        try:
            D[name] = float(stuff[Drow][n])
        except ValueError:
            D[name] = stuff[Drow][n]

        # define the constant parameters (names are C parameters, values are D parameters)
        if type(C[name]) is str:
            const[C[name]] = D[name]

    if version != '':
        try:
            [head, ver] = version.split(' ')  # split at space
            version = ver
        except ValueError:
            pass

    return names, units, data, A, B, C, D, const, version

#######################################################################


def load_header(Inputpath):
    # function loads in header from raw time series data csv input file or header input file. Stores variable names,
    # units, header parameters in dictionaries
    # same as load_timeseries_with_header() but without data series
    # Input: Inputpath: csv file to load. example: Data/alcohol/alcohol_test1/alcohol_test1_RawData.csv
    
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    A = {}  # dictionary keys are variable names, values are A parameters (span)
    B = {}  # dictionary keys are variable names, values are B parameters (offset)
    C = {}  # dictionary keys are variable names, values are C parameters (constant variable names)
    D = {}  # dictionary keys are variable names, values are D parameters (constant variable values)
    const = {}  # dictionary keys are constant variable names(C parameters), values are constant variable values
    # (D parameters)
    
    # load input file
    stuff = []
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    # find the row indices for the header and the data
    for n, row in enumerate(stuff[:100]):  # iterate through first 101 rows to look for the header
        if row[0] == '#A:':
            Arow = n
        if row[0] == '#B:':
            Brow = n 
        if row[0] == '#C:':
            Crow = n
        if row[0] == '#D:':
            Drow = n      
        if row[0] == '#units:':
            unitsrow = n
        if row[0] == 'time':
            namesrow = n
        
    names = stuff[namesrow]
    for n, name in enumerate(names):
        units[name] = stuff[unitsrow][n]
        try:
            A[name] = float(stuff[Arow][n])
        except ValueError:
            A[name] = stuff[Arow][n]
        try:
            B[name] = float(stuff[Brow][n])
        except ValueError:
            B[name] = stuff[Brow][n]
        try:
            C[name] = float(stuff[Crow][n])
        except ValueError:
            C[name] = stuff[Crow][n]
        try:
            D[name] = float(stuff[Drow][n])
        except ValueError:
            D[name] = stuff[Drow][n]

        # define the constant parameters (names are C parameters, values are D parameters)
        if type(C[name]) is str:
            const[C[name]] = D[name]   
         
    return names, units, A, B, C, D, const
    
##########################################################################################


def load_timeseries(Inputpath):
    # function loads in time series data from csv input file and stores variable names, units, and time series in
    # dictionaries
    # Input: Inputpath: csv file to load. example: Data/alcohol/alcohol_test1/alcohol_test1_RawData.csv
    
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    data = {}  # dictionary keys are variable names, values are time series as a list
    
    # load input file
    stuff = []
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)
        
    names = stuff[0]  # first row is channel names
    for n, name in enumerate(names):
        units[name] = stuff[1][n]  # second row is units
        data[name] = [x[n] for x in stuff[2:]]  # data series
        for m, val in enumerate(data[name]):
            try: 
                data[name][m] = float(data[name][m])
            except ValueError:
                pass

    return names, units, data
#######################################################################


def load_L2_constant_inputs(Inputpath):
    # function loads in variables from csv input file and stores variable names, units, and values in dictionaries
    # Input: Inputpath: csv file to load. example:
    # C:\Mountain Air\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_EnergyInputs.csv
    names = []  # list of variable names
    units = {}  # dictionary keys are variable names, values are units
    val = {}  # dictionary keys are variable names, values are variable values
    data = {}  # Dictionary with nested dictionaries defined below
    average = {}  # Dictionary of value averages, key is names
    N = {}  # Dictionary of number of data points averages, key is names
    stdev = {}  # Dictionary of standard deviation of values, key is names
    Interval = {}  # Dictionary of 95% confidence value, key is names
    High = {}  # Dictionary of high tier estimate based on 95% confidence, key is names
    Low = {}  # Dictionary of low tier estimate based on 95% confidence, key is names
    COV = {}  # Dictionary of coifficient of variation, key is names
    CI = {}  # Dictionary of confidence interval (high and low estimates), key is names

    # load input file
    stuff = []
    with open(Inputpath) as f:
        reader = csv.reader(f)
        for row in reader:
            stuff.append(row)

    # Find each row and store the index
    for i, value in enumerate(stuff[0]):
        if stuff[0][i] == 'average':
            averagerow = i
        elif stuff[0][i] == 'N':
            Nrow = i
        elif stuff[0][i] == 'stdev':
            stdevrow = i
        elif stuff[0][i] == 'Interval' or stuff[0][i] == 'interval':
            intervalrow = i
        elif stuff[0][i] == 'High Tier Estimate' or stuff[0][i] == 'high_tier':
            highrow = i
        elif stuff[0][i] == 'Low Tier Estimate' or stuff[0][i] == 'low_tier':
            lowrow = i
        elif stuff[0][i] == 'COV':
            COVrow = i
        elif stuff[0][i] == 'CI':
            CIrow = i

    for row in stuff:  # Grab names
        names.append(row[0])

    n = 0
    for name in names:  # Find values for each column
        try:
            units[name] = stuff[n][1]
        except IndexError:
            units[name] = ''
        try:
            average[name] = stuff[n][averagerow]
        except IndexError:
            average[name] = ''
        try:
            N[name] = stuff[n][Nrow]
        except IndexError:
            N[name] = ''
        try:
            stdev[name] = stuff[n][stdevrow]
        except IndexError:
            stdev[name] = ''
        try:
            Interval[name] = stuff[n][intervalrow]
        except IndexError:
            Interval[name] = ''
        try:
            High[name] = stuff[n][highrow]
        except IndexError:
            High[name] = ''
        try:
            Low[name] = stuff[n][lowrow]
        except IndexError:
            Low[name] = ''
        try:
            COV[name] = stuff[n][COVrow]
        except IndexError:
            COV[name] = ''
        try:
            CI[name] = stuff[n][CIrow]
        except IndexError:
            CI[name] = ''
        try:
            val[name] = stuff[n][2:averagerow]  # Values is a list of values
        except IndexError:
            val[name] = ['']
        n += 1

    # create nested dictionary
    data['average'] = average
    data['N'] = N
    data['stdev'] = stdev
    data['Interval'] = Interval
    data['High Tier'] = High
    data['Low Tier'] = Low
    data['COV'] = COV
    data['CI'] = CI

    return names, units, val, data
#######################################################################


def write_constant_outputs(Outputpath, Names, Units, Val, Unc, Uval):
    # function writes output variables from dictionaries to csv output file
    # Inputs:

    # Outputpath: output csv file that will be created. example:
    # Data/alcohol/alcohol_test1/alcohol_test1_EnergyInputs.csv
    # Names: list of variable names
    # units: dictionary keys are variable names, values are units
    # val: dictionary keys are variable names, values are variable values
    # nom: dictionary keys are variable names, values are variable nominal values
    # unc: dictionary keys are variable names, values are variable uncertainty values
    
    # store data as a list of lists to print by row
    for name in Names:
        try:  # see if a nominal value exists
            Val[name]
        except AttributeError:  # if not then
            try:  # try getting the nominal value from the ufloat
                Val[name] = Uval[name].n
            except AttributeError:  # and if that doesn't work then define the nominal value as the single value
                Val[name] = Uval[name]
        try:  # see if uncertainty value exists
            Unc[name]
        except AttributeError:  # if not then
            try:  # try getting the uncertainty value from the ufloat
                Unc[name] = Uval[name].s
            except AttributeError:
                Unc[name] = ''  # and if that doesn't work then define the uncertainty value as blank
    
    output = []  # initialize list of lines
    for name in Names:  # for each variable
        row = [name, Units[name], Val[name], Unc[name]]  # initialize row
        output.append(row)  # add the row to output list

    # print to the output file
    with open(Outputpath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################


def write_timeseries_with_header(Outputpath, Names, Units, Data, A, B, C, D):
    # function writes time series data csv output file including raw data header with calibration parameters.
    # All variables are taken from dictionaries.
    # Inputs:
    # Outputpath: output csv file that will be created. example:
    # Data/alcohol/alcohol_test1/alcohol_test1_RawData_Recalibrated.csv
    # Names: list of variable names
    # Units: dictionary keys are channel names, values are units
    # Data: dictionary keys are channel names, values are time series as a list
    # A: dictionary keys are channel names, values are A parameters in header
    # B: dictionary keys are channel names, values are B parameters in header
    # C: dictionary keys are channel names, values are C parameters in header
    # D: dictionary keys are channel names, values are D parameters in header
    
    # make lists for each header line
    Arow = []
    Brow = []
    Crow = []
    Drow = []
    Unitsrow = []  # initialize empty rows for the header
    for name in Names:
        Arow.append(A[name])
        Brow.append(B[name])
        Crow.append(C[name])
        Drow.append(D[name])
        Unitsrow.append(Units[name])
        
    # store data as a list of lists to print by row
    output = [Arow, Brow, Crow, Drow, Unitsrow, Names]  # initialize list of output lines starting with header
    for n, val in enumerate(Data['time']):  # for each data point in the time series
        row = []  # initialize blank row
        for name in Names:  # for each channel
            row.append(Data[name][n])  # add the data point
        output.append(row)  # add the row to the output list
    
    # print to the output file
    with open(Outputpath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################


def write_header(Outputpath, Names, Units, A, B, C, D):
    # function writes raw data header to csv file. Same as write_timeseries_with_header() but without data series.
    # All variables are taken from dictionaries.
    # Inputs:
    # Outputpath: output csv file that will be created. example:
    # Data/alcohol/alcohol_test1/alcohol_test1_RawData_Recalibrated.csv
    # Names: list of variable names
    # Units: dictionary keys are channel names, values are units
    # A: dictionary keys are channel names, values are A parameters in header
    # B: dictionary keys are channel names, values are B parameters in header
    # C: dictionary keys are channel names, values are C parameters in header
    # D: dictionary keys are channel names, values are D parameters in header
    
    # make lists for each header line
    Arow = ['#A:']
    Brow = ['#B:']
    Crow = ['#C:']
    Drow = ['#D:']
    Unitsrow = ['#units:']  # initialize empty rows for the header
    for name in Names:
        Arow.append(A[name])
        Brow.append(B[name])
        Crow.append(C[name])
        Drow.append(D[name])
        Unitsrow.append(Units[name])

    if 'time' not in Names:
        Names.insert(0, 'time')
    # store data as a list of lists to print by row
    output = [Arow, Brow, Crow, Drow, Unitsrow, Names]  # initialize list of output lines starting with header
    
    # print to the output file
    with open(Outputpath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################


def write_timeseries(Outputpath, Names, Units, Data):
    # function writes time series data csv output file. All variables are taken from dictionaries.
    # Inputs:
    # Outputpath: output csv file that will be created. example:
    # Data/alcohol/alcohol_test1/alcohol_test1_RawData_Recalibrated.csv
    # Names: list of variable names
    # Units: dictionary keys are channel names, values are units
    # Data: dictionary keys are channel names, values are time series as a list
    
    # make list for units row

    Unitsrow = []  # initialize empty row
    for name in Names:
        Unitsrow.append(Units[name])
        
    # store data as a list of lists to print by row
    output = [Names, Unitsrow]  # initialize list of output lines starting with header
    try:
        for n, val in enumerate(Data['time']):  # for each data point in the time series
            row = []  # initialize blank row
            for name in Names:  # for each channel
                row.append(Data[name][n])  # add the data point
            output.append(row)  # add the row to the output list
    except KeyError:
        for n, val in enumerate(Data['time_test']):
            row = []  # initialize blank row
            for name in Names:  # for each channel
                row.append(Data[name][n])  # add the data point
            output.append(row)  # add the row to the output list
            
    # print to the output file
    with open(Outputpath, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################


def write_timeseries_with_uncertainty(Outputpath, Names, Units, Data):
    # function writes time series data csv output file. All variables are taken from dictionaries.
    # Inputs:
    # Outputpath: output csv file that will be created. example:
    # C:\Mountain Air\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
    # Names: list of variable names
    # Units: dictionary keys are channel names, values are units
    # Data: dictionary keys are channel names, values are time series as a list
    timestampobject = dt.now()  # get timestamp from operating system
    timestampstring = timestampobject.strftime("%H:%M:%S")
    print('start write_timeseries_with_uncertainty ' + timestampstring)
    newnames = []
    regex1 = re.compile('_smooth')
    regex2 = re.compile('_Ave')
    for name in Names:
        if re.search(regex1, name) or re.search(regex2, name):
            newnames.append(name)
            newnames.append(name + '_uc')
            Units[name + '_uc'] = Units[name]
            try:  # ufloats
                Data[name + '_uc'] = [''] * len(Data[name])
                Data[name] = unumpy.nominal_values(Data[name])
            except AttributeError:  # not ufloats
                Data[name + '_uc'] = [''] * len(Data[name])
            print("UC not printed ", name, ", too slow")
        else:
            newnames.append(name)
            newnames.append(name + '_uc')
            Units[name + '_uc'] = Units[name]
            try:  # ufloats
                Data[name + '_uc'] = unumpy.std_devs(Data[name])
                Data[name] = unumpy.nominal_values(Data[name])
            except AttributeError:  # not ufloats
                Data[name + '_uc'] = [''] * len(Data[name])
            print(name)

    timestampobject = dt.now()  # get timestamp from operating system
    timestampstring = timestampobject.strftime("%H:%M:%S")
    print('updated Data ' + timestampstring)

    write_timeseries(Outputpath, newnames, Units, Data)
########################################################################


def write_timeseries_without_uncertainty(Outputpath, Names, Units, Data):
    # similar to write_timeseries_with_uncertainty but only takes the nominal value
    # function writes time series data csv output file. All variables are taken from dictionaries.
    # Inputs:
    # Outputpath: output csv file that will be created. example:
    # C:\Mountain Air\DataProcessing\LEMS\LEMS-Data-Processing\Data\CrappieCooker\CrappieCooker_RawDataOutput.csv
    # Names: list of variable names
    # Units: dictionary keys are channel names, values are units
    # Data: dictionary keys are channel names, values are time series as a list

    timestampobject = dt.now()  # get timestamp from operating system
    timestampstring = timestampobject.strftime("%H:%M:%S")
    print('start write_timeseries_without_uncertainty ' + timestampstring)

    # check for the file
    if os.path.isfile(Outputpath):
        os.remove(Outputpath)  # and remove it (because writing appends)

    # make list for names row and units row
    Namesrow = []  # initialize empty row
    Unitsrow = []  # initialize empty row
    for name in Names:
        Namesrow.append(name)
        # Namesrow.append(name+'_uc')
        Unitsrow.append(Units[name])
        # Unitsrow.append(Units[name])
    # store data as a list of lists to print by row
    output = [Namesrow, Unitsrow]  # initialize list of output lines starting with header
    for n in range(len(Data['time'])):  # for each data point in the time series
        row = []  # initialize blank row
        for name in Names:  # for each channel
            try:  # ufloat
                row.append(Data[name][n].n)  # add the data point
                # row.append(Data[name][n].s)          #add the data point
            except AttributeError:  # not ufloat
                row.append(Data[name][n])  # add the data point
                # row.append('')                                  #add the data point
        output.append(row)  # add the row to the output list
        if n % 1000 == 0 or n == len(Data['time']) - 1:
            # print to the output file
            with open(Outputpath, 'a', newline='') as csvfile:
                writer = csv.writer(csvfile)
                for outrow in output:
                    writer.writerow(outrow)
            output = []

            timestampobject = dt.now()  # get timestamp from operating system
            timestampstring = timestampobject.strftime("%H:%M:%S")
            print(str(n) + ' ' + timestampstring)

    # print to the output file
    with open(Outputpath, 'w' ,newline='') as csvfile:
        writer = csv.writer(csvfile)
        for row in output:
            writer.writerow(row)
########################################################################


def write_logfile(Logpath, Logs):
    # writes to logfile.txt to document data manipulations
    # Inputs:
    # Logpath: logfile.txt path. example:  Data/alcohol/alcohol_test1/alcohol_test1_log.txt
    # Logs: list of lines that will get logged to the file
    with open(Logpath, 'a') as logfile: 
        for log in Logs:
            logfile.write('\n'+log)
#######################################################################


def create_header(headerpath, names, data, logger, logs):
    # Function purpose: Ask user for calibration parameters to write a header and or recalibrate data. Used when a
    # header is not created in the raw data output. User can chose to not create a header or not recalibrate data

    # Inputs:
    # headerpath: file path to write header to
    # names: list of variable names that will get calibration parameters
    # data: dictionary of timeseries data with names as key that will get recalibrated
    # logger: Python logging function
    # logs: list of important events

    # Outputs:
    # header with current A and B calibration parameters (if created)

    # Return:
    # Recalibrated data (or input data if not recalibrated)
    # Logs: list of important events

    # function called by PEMS_2041, LEMS_3001, LEMS_3002, LEMS3015

    # Record start time of script
    func_start_time = dt.now()
    log = f"Started at: {func_start_time}"
    print(log)
    logger.info(log)
    logs.append(log)

    # Log script version if available
    try:
        version = subprocess.check_output(
            ["git", "log", "-n", "1", "--pretty=format:%h", "--", __file__], text=True
        ).strip()
    except subprocess.CalledProcessError:
        version = "unknown_version"
    log = f"Version: {version}"
    print(log)
    logger.info(log)
    logs.append(log)

    # Create dictionaries for old and new parameters keys are names
    A_old, B_old, A_new, B_new = {}, {}, {}, {}

    # Go x names at a time to not make the message box too long
    batch_size = 4
    for i in range(0, len(names), batch_size):
        # select the next batch of names
        batch_names = names[i:i + batch_size]

        # Prepare fields for all names, appending _A_old, _B_old, _A_new, _B_new suffixes
        fields = []
        for name in batch_names:
            fields.append(f"{name}_A_old")
            fields.append(f"{name}_A_new")
            fields.append(f"{name}_B_old")
            fields.append(f"{name}_B_new")

        msg = f"If recalibration is desired, enter the current and new A and B calibration constants for the desired" \
              f"fields. To skip a calibration, leave the entry blank. To skip all calibration, click 'cancel'.\n" \
              f"To record current calibration constants but not recalibrate (sensor box has already used current " \
              f"calibration constants internally), just enter new calibration constants and no old ones."

        title = "Calibration Constants"
        values = easygui.multenterbox(msg, title, fields)

        # Check if the user pressed cancel or left fields blank, and assign NaN accordingly
        if values is None:
            write = False  # don't write a header
            # If the user cancels, set all values to NaN
            for name in names:
                A_old[name] = B_old[name] = A_new[name] = B_new[name] = math.nan
        else:
            write = True  # write a header

            i = 0
            # Current parameters (A_old, B_old)
            for j, name in enumerate(batch_names):
                try:
                    A_old[name] = float(values[j + i]) if values[j + i] else math.nan
                except (ValueError, IndexError):
                    A_old[name] = math.nan
                try:
                    B_old[name] = float(values[j + i + 2]) if values[j + i + 2] else math.nan
                except (ValueError, IndexError):
                    B_old[name] = math.nan
                try:
                    A_new[name] = float(values[j + i + 1]) if values[j + i + 1] else math.nan
                except (ValueError, IndexError):
                    A_new[name] = math.nan
                try:
                    B_new[name] = float(values[j + i + 3]) if values[j + i + 3] else math.nan
                except (ValueError, IndexError):
                    B_new[name] = math.nan

                i += 4

                # Validate that if one old value is entered, both are required
                if (not math.isnan(A_old[name])) or (not math.isnan(B_old[name])):  # if either A_old or B_old entered
                    if math.isnan(A_old[name]) or math.isnan(B_old[name]):  # if any value is blank
                        msg = "Both A and B values must be entered or left blank to function properly"
                        title = "Input Required"
                        fields = [f"{name}_A_old", f"{name}_B_old"]
                        default = [A_old[name], B_old[name]]
                        values = easygui.multenterbox(msg, title, fields, default)  # reprompt
                        try:
                            A_old[name] = float(values[0])
                        except (ValueError, IndexError):
                            A_old[name] = math.nan
                        try:
                            B_old[name] = float(values[1])
                        except (ValueError, IndexError):
                            B_old[name] = math.nan

                # Validate that if one new value is entered, both are required
                if (not math.isnan(A_new[name])) or (not math.isnan(B_new[name])):  # if either A_old or B_old entered
                    if math.isnan(A_new[name]) or math.isnan(B_new[name]):  # if any value is blank
                        msg = "Both A and B values must be entered or left blank to function properly"
                        title = "Input Required"
                        fields = [f"{name}_A_new", f"{name}_B_new"]
                        default = [A_new[name], B_new[name]]
                        values = easygui.multenterbox(msg, title, fields, default)  # reprompt
                        try:
                            A_new[name] = float(values[0])
                        except (ValueError, IndexError):
                            A_new[name] = math.nan
                        try:
                            B_new[name] = float(values[1])
                        except (ValueError, IndexError):
                            B_new[name] = math.nan

    if write:  # write a header using the new inputs
        C_new = {}  # blank values, not using C
        for name in names:
            C_new[name] = math.nan
        D_new = {}  # blank values, not using D
        for name in names:
            D_new[name] = math.nan
        units = {}  # blank values for log units
        for name in names:
            units[name] = ''
        write_header(headerpath, names, units, A_new, B_new, C_new, D_new)

        line = f"Created: {headerpath} using entered A and B calibration parameters"
        print(line)
        logger.info(line)
        logs.append(line)

    const_old = {}  # left blank, not used in redofirmware currently
    const_new = {}
    if 'time' in names:
        names.remove('time')

    all_nan = True
    for name in A_old:
        if not math.isnan(A_old[name]) or not math.isnan(B_old[name]):
            all_nan = False

    if all_nan:
        # redo firmware calculations
        [data_new, add_logs] = RedoFirmwareCalcs(names, A_old, B_old, const_old, data,
                                                 A_new, B_new, const_new, logger)
        logs = logs + add_logs

        data = data_new

    end_time = dt.now()
    log = f"Execution time: {end_time - func_start_time}"
    print(log)
    logger.info(log)
    logs.append(log)

    return data, logs
