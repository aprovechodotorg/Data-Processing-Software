import csv
from openpyxl import load_workbook
import LEMS_DataProcessing_IO as io
from datetime import datetime as dt

def LEMS_CustomFormatted_L3(inputpath, outputpath, outputexcel, csvpath, logpath):
    """
    Reads data from a source CSV, maps it to a template based on data_keys,
    and writes to new CSV/Excel files.
    """

    #Function intakes list of inputpaths and creates comparison between values in list.
    ver = '0.0'

    timestampobject = dt.now()  # get timestamp from operating system for log file
    timestampstring = timestampobject.strftime("%Y%m%d %H:%M:%S")

    line = 'LEMS_CustomFormatted_L3 v' + ver + '   ' + timestampstring  # Add to log
    print(line)
    logs = [line]

    # 1. Load Data and Units from CustomCutTable_L3
    # Dictionary structure: source_data[variable] = {'units': '...', 'values': {name_key: value}}
    source_data = {}
    with open(inputpath, 'r') as f:
        reader = list(csv.reader(f))
        name_keys = reader[2]  # Row with 1B, 1R1, etc.

        for row in reader[2:]:
            if not row or len(row) < 2: continue
            var_name = row[0]
            units = row[1]
            source_data[var_name] = {'units': units, 'values': {}}

            for i in range(2, len(row)):
                key = name_keys[i]
                if key:
                    source_data[var_name]['values'][key] = row[i]
    line = 'loaded input data: ' + inputpath
    print(line)
    logs.append(line)

    # 2. Load Template and Map Data
    # We use the openpyxl library to handle the .xlsx template formatting
    wb = load_workbook(csvpath)

    line = 'loaded output template: ' + csvpath
    print(line)
    logs.append(line)

    sheet = wb.active

    # Find the header row (containing name_keys) and the data_key columns
    header_row_idx = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith('data_key'):
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break

    data_key_cols = []
    if header_row_idx is not None:
        for cell in sheet[header_row_idx]:
            if cell.value and str(cell.value).startswith('data_key'):
                data_key_cols.append(cell.column)
        data_key_cols.sort()

    # Map each data_key column to its units column and its template_name_keys
    data_key_configs = []
    for idx, dk_col in enumerate(data_key_cols):
        # Determine units_col: check if cell to the left in header row contains 'units'
        units_col = None
        if dk_col > 1:
            left_val = sheet.cell(row=header_row_idx, column=dk_col - 1).value
            if left_val and str(left_val).strip().lower() == 'units':
                units_col = dk_col - 1

        # Determine the columns to the right adhering to this data_key
        # The next data_key column (if any) sets the boundary
        next_dk_col = data_key_cols[idx + 1] if idx + 1 < len(data_key_cols) else float('inf')

        t_name_keys = {}
        for col_cell in sheet[header_row_idx]:
            c_col = col_cell.column
            # Only include columns strictly between dk_col and next_dk_col
            if dk_col < c_col < next_dk_col and col_cell.value:
                val_str = str(col_cell.value).strip()
                if val_str.lower() not in ('units',) and not val_str.startswith('data_key'):
                    t_name_keys[val_str] = c_col

        data_key_configs.append({
            'data_key_col': dk_col,
            'units_col': units_col,
            'template_name_keys': t_name_keys
        })

    # 3. Fill the template with data and units
    for row_cells in sheet.iter_rows(min_row=1):
        row_idx = row_cells[0].row
        if row_idx == header_row_idx:
            continue

        for config in data_key_configs:
            dk_col = config['data_key_col']
            units_col = config['units_col']
            t_name_keys = config['template_name_keys']

            d_key_val = sheet.cell(row=row_idx, column=dk_col).value

            # Resolve d_key against source_data keys
            d_key = None
            if d_key_val is not None:
                if d_key_val in source_data:
                    d_key = d_key_val
                elif str(d_key_val).strip() in source_data:
                    d_key = str(d_key_val).strip()

            if d_key is not None:
                # Write Units
                if units_col is not None:
                    sheet.cell(row=row_idx, column=units_col).value = source_data[d_key]['units']

                # Write Values
                for n_key, col_idx in t_name_keys.items():
                    # Check both original key and stripped key
                    val = None
                    if n_key in source_data[d_key]['values']:
                        val = source_data[d_key]['values'][n_key]
                    elif n_key.strip() in source_data[d_key]['values']:
                        val = source_data[d_key]['values'][n_key.strip()]

                    if val is not None:
                        try:
                            # Clean string and convert to float if numeric
                            if str(val).strip() == '' or str(val).lower() == 'nan':
                                sheet.cell(row=row_idx, column=col_idx).value = None
                            else:
                                sheet.cell(row=row_idx, column=col_idx).value = round(float(val), 2)
                        except (ValueError, TypeError):
                            sheet.cell(row=row_idx, column=col_idx).value = val

    # 4. Save Outputs
    # Save Excel version
    wb.save(outputexcel)
    # Log the action
    line = 'Formatted Custom Cut Table created: ' + outputexcel
    print(line)
    logs.append(line)

    # Save CSV version (using logic from your write_timeseries function)
    with open(outputpath, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

    # Log the action
    line = 'Formatted Custom Cut Table created: ' + outputpath
    print(line)
    logs.append(line)

    # print to log file
    io.write_logfile(logpath, logs)