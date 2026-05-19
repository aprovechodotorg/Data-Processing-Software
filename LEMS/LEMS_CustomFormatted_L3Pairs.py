import csv
from openpyxl import load_workbook
import LEMS_DataProcessing_IO as io
from datetime import datetime as dt


def LEMS_CustomFormatted_L3Pairs(inputpath, outputpath, outputexcel, template, logpath):
    """
    Reads data from a source CSV, maps it to a template based on data_keys,
    and writes to new CSV/Excel files.
    """

    ver = '1.1'

    timestampobject = dt.now()  # get timestamp from operating system for log file
    timestampstring = timestampobject.strftime("%Y%m%d %H:%M:%S")

    line = 'LEMS_CustomFormatted_L3Pairs v' + ver + '   ' + timestampstring  # Add to log
    print(line)
    logs = [line]

    # 1. Load Data and Units from CustomCutTable_L3
    # Dictionary structure: source_data[variable] = {'units': '...', 'values': {name_key: value}}

    # Read from row 1 onwards to capture all data records
    source_data = {}
    with open(inputpath, 'r') as f:
        reader = list(csv.reader(f))
        name_keys = reader[0]  # Row with 1B, 1R1, etc.

        for row in reader[1:]:
            if not row or len(row) < 2: continue
            var_name = row[0]
            units = row[1]
            source_data[var_name] = {'units': units, 'values': {}}
            #don't copy the confidence interval or N stats since only want to look at percent difference which doesn't have CI or N the same way
            if var_name == "confidence":
                break

            for i in range(2, len(row)):
                key = name_keys[i]
                if key:
                    source_data[var_name]['values'][key] = row[i]

    line = 'loaded input data: ' + inputpath
    print(line)
    logs.append(line)

    # 2. Load Template and Map Data
    wb = load_workbook(template)

    line = 'loaded output template: ' + template
    print(line)
    logs.append(line)

    sheet = wb.active

    # Find the header row (containing name_keys) and the data_key column
    template_name_keys = {}  # Mapping of name_key -> column index
    data_key_col = None
    units_col = None

    # Search for the header row and data_key column
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'data_key':
                data_key_col = cell.column
                units_col = cell.column - 1  # Units are one col to the left
                # Map the experiment IDs (d1B/1R1...) in this row to their columns
                for col_cell in sheet[cell.row]:
                    if col_cell.column > data_key_col and col_cell.value:
                        template_name_keys[str(col_cell.value).strip()] = col_cell.column
                break
        if data_key_col:
            break

    if not data_key_col:
        line = "Error: 'data_key' identifier column not found in template."
        print(line)
        logs.append(line)
        io.write_logfile(logpath, logs)
        return

    # 3. Fill the template with data and units
    for row_cells in sheet.iter_rows(min_row=1):
        row_idx = row_cells[0].row
        d_key = sheet.cell(row=row_idx, column=data_key_col).value

        if d_key:
            d_key = str(d_key).strip()

        if d_key in source_data:
            # Write Units
            if units_col > 0:
                sheet.cell(row=row_idx, column=units_col).value = source_data[d_key]['units']

            # Write Values
            for n_key, col_idx in template_name_keys.items():
                # FIX 2: Implement partial matching because template column names (e.g. 'd1B/1R1')
                # are substrings of the CSV headers (e.g. 'd1B/1R1 % Change')
                target_csv_key = None
                if n_key in source_data[d_key]['values']:
                    target_csv_key = n_key
                else:
                    for csv_key in source_data[d_key]['values']:
                        if n_key in csv_key:
                            target_csv_key = csv_key
                            break

                # If a matching column header was found, extract and write the value
                if target_csv_key:
                    val = source_data[d_key]['values'][target_csv_key]
                    try:
                        # Clean string and convert to float if numeric
                        if str(val).strip() == '' or str(val).lower() == 'nan':
                            sheet.cell(row=row_idx, column=col_idx).value = None
                        else:
                            sheet.cell(row=row_idx, column=col_idx).value = round(float(val),1)
                    except (ValueError, TypeError):
                        sheet.cell(row=row_idx, column=col_idx).value = val

    # 4. Save Outputs
    # Save Excel version
    wb.save(outputexcel)
    line = 'Formatted Custom Cut Table created: ' + outputexcel
    print(line)
    logs.append(line)

    # Save CSV version
    with open(outputpath, 'w', newline='') as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

    line = 'Formatted Custom Cut Table created: ' + outputpath
    print(line)
    logs.append(line)

    # print to log file
    io.write_logfile(logpath, logs)