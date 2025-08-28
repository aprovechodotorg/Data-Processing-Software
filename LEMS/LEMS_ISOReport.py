import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import math
import  LEMS_DataProcessing_IO as io


def LEMS_ISOReport(data_values, units, outputpath, logpath):
    """
    Creates a formatted Excel table showing metrics similar to the reference image.

    Parameters:
    data_values (dict): Dictionary containing metrics data from PEMS_L2 function
    output_file_path (str): Path for saving the output Excel file
    """

    logs = []

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ISO Report Table"

    # Define border styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Define fills
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Setup the headers
    ws.merge_cells('A1:B2')
    ws['A1'] = "Metric"
    ws['A1'].fill = header_fill

    ws.merge_cells('C1:F1')
    ws['C1'] = "Test Sequence Phase"
    ws['C1'].fill = header_fill

    ws['C2'] = "High"
    ws['C2'].fill = header_fill
    ws['D2'] = "Medium"
    ws['D2'].fill = header_fill
    ws['E2'] = "Low"
    ws['E2'].fill = header_fill
    ws['F2'] = "Combined"
    ws['F2'].fill = header_fill

    ws.merge_cells('G1:G2')
    ws['G1'] = "Tier Rating"
    ws['G1'].fill = header_fill
    ws['G1'].alignment = Alignment(vertical='center')
    ws['G1'].font = Font(bold=True)

    # Set column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12  # Width for the label column
    for col in range(3, 8):  # Adjusted for the new column
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Define the metrics to be included in the table
    metrics = [
        {"name": f"Thermal Efficiency without char ({units.get('eff_wo_char_hp', 'N/A')})", "key_base": "eff_wo_char",
         "tier_key": "tier_eff_wo_char"},
        {"name": f"Thermal Efficiency with char ({units.get('eff_w_char_hp', 'N/A')})", "key_base": "eff_w_char",
         "tier_key":"tier_eff_w_char"},
        {"name": f"Char Energy Productivity ({units.get('char_energy_productivity_hp', 'N/A')})", "key_base":
            "char_energy_productivity", "tier_key": None},
        {"name": f"Char Mass Productivity ({units.get('char_mass_productivity_hp', 'N/A')})", "key_base":
            "char_mass_productivity", "tier_key": None},
        {"name": f"Cooking Power ({units.get('cooking_power_hp', 'N/A')})", "key_base": "cooking_power",
         "tier_key": None},
        {"name": f"Fuel Burning Rate ({units.get('burn_rate_dry_hp', 'N/A')})", "key_base": "burn_rate_dry",
         "tier_key": None},
        {"name": f"PM2.5 per useful energy ({units.get('PM_useful_eng_deliver_hp', 'N/A')})", "key_base":
            "PM_useful_eng_deliver", "tier_key": "tier_PM_useful_eng_deliver"},
        {"name": f"CO per useful energy ({units.get('CO_useful_eng_deliver_hp', 'N/A')})", "key_base":
            "CO_useful_eng_deliver", "tier_key": "tier_CO_useful_eng_deliver"},
        {"name": "Safety", "key_base": None, "tier_key": None},
        {"name": "Durability", "key_base": None, "tier_key": None}
    ]

    # Current row index (starting after headers)
    row_idx = 3

    # Phases to include in the table
    phases = ["_hp", "_mp", "_lp", "_weighted"]
    phase_cols = {'_hp': 'C', '_mp': 'D', '_lp': 'E', '_weighted': 'F'}  # Adjusted for the new column

    # Process each metric
    for metric in metrics:
        # For Safety and Durability metrics (simplified case)
        #if metric["key_base"] is None:
            #ws[f'A{row_idx}'] = metric["name"]
            #ws[f'A{row_idx}'].alignment = Alignment(wrap_text=True, horizontal='center')
            #ws[f'B{row_idx}'] = "Score"
            #ws[f'B{row_idx}'].alignment = Alignment(horizontal='center')
            #for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:  # Adjusted for the new column
                #ws[f'{col}{row_idx}'].border = thin_border
            #ws[f'G{row_idx}'] = "N/A"
            #row_idx += 1
            #continue

        # Decide if this metric should have 2 or 3 rows
        if metric["key_base"] == "char_energy_productivity" or metric["key_base"] == "char_mass_productivity" or metric["key_base"] == "cooking_power" or metric["key_base"] == "burn_rate_dry":
            labels = ["Mean", "SD"]
        elif metric["key_base"] is None:
            labels = ["Score"]
        else:
            labels = ["Mean", "SD", "90% CI"]

        num_rows = len(labels)

        # For regular metrics with multiple rows
        # First, write the metric name
        ws[f'A{row_idx}'] = metric["name"]

        # Create merge after writing the value
        merge_range = f'A{row_idx}:A{row_idx + num_rows - 1}'
        ws.merge_cells(merge_range)
        ws[f'A{row_idx}'].alignment = Alignment(vertical='center', wrap_text=True)
        ws[f'A{row_idx}'].border = thin_border

        # Set up tier rating cell
        #if metric["tier_key"] and metric["tier_key"] in data_values:
            #ws[f'G{row_idx}'] = data_values[metric["tier_key"]]["average"]
        #else:
            #ws[f'G{row_idx}'] = "N/A"

        #merge_range = f'G{row_idx}:G{row_idx + num_rows - 1}'
        #ws.merge_cells(merge_range)
        #for i in range(num_rows):
            #ws[f'G{row_idx + i}'].border = thin_border
        #ws[f'G{row_idx}'].alignment = Alignment(vertical='center', horizontal='center')

        # Special handling for Cooking Power - merge Combined column and Tier column
        if metric["key_base"] == "cooking_power" or metric["key_base"] == "burn_rate_dry":
            # Merge the Combined column cells
            merge_range_combined = f'F{row_idx}:G{row_idx + num_rows - 1}'
            ws.merge_cells(merge_range_combined)
            ws[f'F{row_idx}'] = "N/A"
            ws[f'F{row_idx}'].alignment = Alignment(vertical='center', horizontal='center')
            ws[f'F{row_idx}'].border = thin_border
            ws[f'G{row_idx}'].border = thin_border
            ws[f'G{row_idx + 1}'].border = thin_border

            # Merge the Tier column cells and set N/A
            #merge_range_tier = f'G{row_idx}:G{row_idx + num_rows - 1}'
            #ws.merge_cells(merge_range_tier)
            #ws[f'G{row_idx}'] = "N/A"
            #ws[f'G{row_idx}'].alignment = Alignment(vertical='center', horizontal='center')
            #ws[f'G{row_idx}'].border = thin_border

            # Process each phase except weighted for cooking power and burn rate
            for i, label in enumerate(labels):
                curr_row = row_idx + i

                # Add row label
                ws[f'B{curr_row}'] = label
                ws[f'B{curr_row}'].alignment = Alignment(horizontal='center')
                ws[f'B{curr_row}'].border = thin_border

                # Process only the High, Medium, and Low phases
                for phase in phases[:-1]:  # Exclude the weighted phase
                    key = f"{metric['key_base']}{phase}"
                    col = phase_cols[phase]

                    # Handle Mean row
                    if label == "Mean":  # Mean
                        if key in data_values and "average" in data_values[key]:
                            value = data_values[key]["average"]
                            ws[f'{col}{curr_row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(
                                value) else "-"
                        else:
                            ws[f'{col}{curr_row}'] = "-"

                    # Handle SD row
                    elif label == "SD":  # SD
                        if key in data_values and "stdev" in data_values[key]:
                            value = data_values[key]["stdev"]
                            ws[f'{col}{curr_row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(
                                value) else "-"
                        else:
                            ws[f'{col}{curr_row}'] = "-"

                    # Apply borders to the cells
                    ws[f'{col}{curr_row}'].border = thin_border

        else:
            # Set up tier rating cell for other metrics
            if metric["tier_key"] and metric["tier_key"] in data_values:
                try:
                    tier, value = data_values[metric["tier_key"]]["average"].split(" ")
                except:
                    value = data_values[metric["tier_key"]]["average"]
                ws[f'G{row_idx}'] = value
            else:
                ws[f'G{row_idx}'] = "N/A"

            merge_range = f'G{row_idx}:G{row_idx + num_rows - 1}'
            ws.merge_cells(merge_range)
            for i in range(num_rows):
                ws[f'G{row_idx + i}'].border = thin_border
            ws[f'G{row_idx}'].alignment = Alignment(vertical='center', horizontal='center')

            for i, label in enumerate(labels):
                curr_row = row_idx + i

                # Add row label
                ws[f'B{curr_row}'] = label
                ws[f'B{curr_row}'].alignment = Alignment(horizontal='center')
                ws[f'B{curr_row}'].border = thin_border

                # Process each phase
                for phase in phases:
                    key = f"{metric['key_base']}{phase}"
                    col = phase_cols[phase]

                    # Handle Mean row
                    if label == "Mean":  # Mean
                        if key in data_values and "average" in data_values[key]:
                            value = data_values[key]["average"]
                            ws[f'{col}{curr_row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(
                                value) else "-"
                        else:
                            ws[f'{col}{curr_row}'] = "-"

                    # Handle SD row
                    elif label == "SD":  # SD
                        if key in data_values and "stdev" in data_values[key]:
                            value = data_values[key]["stdev"]
                            ws[f'{col}{curr_row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(
                                value) else "-"
                        else:
                            ws[f'{col}{curr_row}'] = "-"

                    # Handle 90% CI row
                    elif label == "90% CI":  # 90% CI
                        if key in data_values and "high_tier" in data_values[key] and "low_tier" in data_values[key]:
                            high = data_values[key]["high_tier"]
                            low = data_values[key]["low_tier"]

                            if not isinstance(high, str) and not math.isnan(high) and not isinstance(low,
                                                                                                     str) and not math.isnan(
                                low):
                                ws[f'{col}{curr_row}'] = f"{round(low, 1)} - {round(high, 1)}"
                                ws[f'{col}{curr_row}'].alignment = Alignment(horizontal='right')
                            else:
                                ws[f'{col}{curr_row}'] = "-"
                        else:
                            ws[f'{col}{curr_row}'] = "-"

                    # Apply borders to the cells
                    ws[f'{col}{curr_row}'].border = thin_border

                if label == 'Score':
                    merge_range_combined = f'C{row_idx}:F{row_idx}'
                    ws.merge_cells(merge_range_combined)
                    ws[f'C{row_idx}'].alignment = Alignment(vertical='center', horizontal='center')
                    ws[f'C{row_idx}'].border = thin_border

            # Move to the next metric (after the rows for this metric)
        row_idx += num_rows

    # Apply borders and alignment to all header cells
    for cell in ['A1', 'B1', 'G1', 'G2', 'C1', 'C2', 'D2', 'E2', 'F2']:
        ws[cell].border = thin_border
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].font = Font(bold=True)

    ######################################################################################
    phases = ['hp', 'mp', 'lp']

    for phase in phases:
        # create a new tab
        if phase == 'hp':
            ws = wb.create_sheet(title="High Power")
        elif phase == 'mp':
            ws = wb.create_sheet(title="Medium Power")
        elif phase == 'lp':
            ws = wb.create_sheet(title="Low Power")

        # define the metrics to be added to the tab
        base_metrics = [
            {"name": "Fuel mass, dry", "key_base": "fuel_dry_mass"},
            {"name": "Fuel moisture content (wet basis)", "key_base": "fuel_mc_1"},
            {"name": "Thermal efficiency with char", "key_base": "eff_w_char"},
            {"name": "Thermal efficiency without char", "key_base": "eff_wo_char"},
            {"name": "Char energy productivity", "key_base": "char_energy_productivity"},
            {"name": "Char mass productivity", "key_base": "char_mass_productivity"},
            {"name": "Fuel burning rate (dry basis)", "key_base": "burn_rate_dry"},
            {"name": "Cooking power", "key_base": "cooking_power"},
            {"name": "Modified combustion efficiency", "key_base": "MCE"},
            {"name": "PM2.5 total mass", "key_base": "PM_total_mass"},
            {"name": "PM2.5 mass per dry fuel mass", "key_base": "PM_fuel_dry_mass"},
            {"name": "PM2.5 mass per fuel energy", "key_base": "PM_fuel_energy_w_char"},
            {"name": "PM2.5 mass per useful energy delivered", "key_base": "PM_useful_eng_deliver"},
            {"name": "PM2.5 mass per time", "key_base": "PM_mass_time"},
            {"name": "CO total mass", "key_base": "CO_total_mass"},
            {"name": "CO mass per dry fuel mass", "key_base": "CO_fuel_dry_mass"},
            {"name": "CO mass per fuel energy", "key_base": "CO_fuel_energy_w_char"},
            {"name": "CO mass per useful energy delivered", "key_base": "CO_useful_eng_deliver"},
            {"name": "CO mass per time", "key_base": "CO_mass_time"},
            {"name": "CO2 total mass", "key_base": "CO2_total_mass"},
            {"name": "CO2 mass per dry fuel mass", "key_base": "CO2_fuel_dry_mass"},
            {"name": "CO2 mass per fuel energy", "key_base": "CO2_fuel_energy_w_char"},
            {"name": "CO2 mass per useful energy delivered", "key_base": "CO2_useful_eng_deliver"},
            {"name": "CO2 mass per time", "key_base": "CO2_mass_time"},
            {"name": "BC total mass", "key_base": "BC_total_mass"},
            {"name": "BC mass per dry fuel mass", "key_base": "BC_fuel_dry_mass"},
            {"name": "BC mass per fuel energy", "key_base": "BC_fuel_energy_w_char"},
            {"name": "BC mass per useful energy delivered", "key_base": "BC_useful_eng_deliver"},
            {"name": "BC mass per time", "key_base": "BC_mass_time"},
        ]

        # Create phase-specific metrics list with keys and units
        metrics = []
        for base_metric in base_metrics:
            key = f"{base_metric['key_base']}_{phase}"

            # Special case for fuel_mc_1 which doesn't have a phase suffix
            if base_metric['key_base'] == "fuel_mc_1":
                key = "fuel_mc_1"

            metrics.append({
                "name": base_metric["name"],
                "key": key,
                "unit": units.get(key, "N/A")  # Use get() to safely handle missing keys
            })

        # Set column widths for High Power tab
        ws.column_dimensions['A'].width = 32  # Metric name
        ws.column_dimensions['B'].width = 10  # Units

        # Determine number of tests based on data
        num_tests = 5  # Default to 5 tests
        # Try to determine actual number of tests from data if available
        for metric in metrics:
            if metric["key"] in data_values and "values" in data_values[metric["key"]]:
                num_tests = len(data_values[metric["key"]]["values"])
                break

        # Set up header row for tab
        ws['A1'] = "Metric"
        ws['B1'] = "Units"
        ws['C1'] = "Mean"
        ws['D1'] = "SD"

        # Add test columns
        for i in range(1, num_tests + 1):
            col_letter = get_column_letter(i + 4)  # E, F, G, etc.
            ws[f'{col_letter}1'] = f"Test {i}"
            ws.column_dimensions[col_letter].width = 10

        # Apply formatting to header row
        for col in range(1, num_tests + 5):  # A through last test column
            cell = ws[f'{get_column_letter(col)}1']
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Fill data for high power tab
        row = 2
        for metric in metrics:
            # Set metric name and unit
            ws[f'A{row}'] = metric["name"]
            ws[f'B{row}'] = metric["unit"]

            # Get data if available
            if metric["key"] in data_values:
                # Mean and SD
                if "average" in data_values[metric["key"]]:
                    value = data_values[metric["key"]]["average"]
                    ws[f'C{row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(value) else "-"
                else:
                    ws[f'C{row}'] = "-"
                ws[f'C{row}'].fill = header_fill

                if "stdev" in data_values[metric["key"]]:
                    value = data_values[metric["key"]]["stdev"]
                    ws[f'D{row}'] = round(value, 1) if not isinstance(value, str) and not math.isnan(value) else "-"
                else:
                    ws[f'D{row}'] = "-"

                # Individual test values
                if "values" in data_values[metric["key"]]:
                    test_values = data_values[metric["key"]]["values"]
                    for i, val in enumerate(test_values):
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                        if i < num_tests:  # Only process up to max number of tests
                            col_letter = get_column_letter(i + 5)  # E, F, G, etc.
                            if not isinstance(val, str) and not math.isnan(val):
                                ws[f'{col_letter}{row}'] = round(val, 1)
                            else:
                                ws[f'{col_letter}{row}'] = "-"
            else:
                # No data available for this metric
                ws[f'C{row}'] = "-"
                ws[f'C{row}'].fill = header_fill
                ws[f'D{row}'] = "-"
                for i in range(num_tests):
                    col_letter = get_column_letter(i + 5)
                    ws[f'{col_letter}{row}'] = "-"

            # Apply borders and alignment to all cells in the row
            for col in range(1, num_tests + 5):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.border = thin_border
                if col >= 3:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

            row += 1

    #########################################################################
    # Create quality control tab
    ws = wb.create_sheet(title="Quality Control")
    subheader_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    # Determine number of tests from data
    num_tests = 1  # Default
    for key in data_values:
        if "values" in data_values[key]:
            test_count = len(data_values[key]["values"])
            if test_count > 0:
                num_tests = test_count
                break

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, num_tests + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # === STOVE INFORMATION SECTION ===

    # Stove Information header row (gray background)
    ws.merge_cells(f'A1:{get_column_letter(num_tests + 2)}1')
    ws['A1'] = "Stove Information"
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True)
    ws['A1'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}1'].border = thin_border

    # Stove Information rows
    info_rows = [
        {"label": "Stove type/model", "key": "stove_type/model"},
        {"label": "Location", "key": "location"},
        {"label": "Fuel species", "key": "fuel_type_1"},
        {"label": "Date", "key": "date"}
    ]

    row_idx = 2
    for info in info_rows:
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws[f'A{row_idx}'] = info["label"]
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='right', vertical='center')

        # Get values from data_values if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""
            if cell_key == "Stove type/model":
                ws[f'{col_letter}{row_idx}'].font = Font(bold=True)

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    # === PM2.5 QUALITY CONTROL SECTION ===
    # Gas Sensor header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "PM2.5 Quality Control"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    # Leak rows
    grav_rows = [
        {"label": "Gravimetric A Leak Rate", "unit": units.get("Gravametric_A_Leak_Rate", 'N/A'),
         "key": "Gravametric_A_Leak_Rate"},
        {"label": "Gravimetric A Leak Check", "unit": "Pass/Fail", "key": "Gravametric_A_Leak_Check"},
        {"label": "Gravimetric B Leak Rate", "unit": units.get("Gravametric_B_Leak_Rate", 'N/A'),
         "key": "Gravametric_B_Leak_Rate"},
        {"label": "Gravimetric B Leak Check", "unit": "Pass/Fail", "key": "Gravametric_B_Leak_Check"}
    ]

    for info in grav_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    # == PM PHASE METRICS ==
    for phase in phases:
        #  header
        ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
        if phase == 'hp':
            p = "High Power"
            ws[f'A{row_idx}'] = "High Power"
        elif phase == 'mp':
            p = "Medium Power"
            ws[f'A{row_idx}'] = "Medium Power"
        elif phase == 'lp':
            p = "Low Power"
            ws[f'A{row_idx}'] = "Low Power"
        ws[f'A{row_idx}'].font = Font(bold=True)
        ws[f'A{row_idx}'].fill = subheader_fill
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row_idx}'].border = thin_border
        for col_idx in range(2, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
        row_idx += 1

        # PM rows
        PM_rows = [
            {"label": f"Mass Scattering Cross Section {p}", "unit": units.get(f"MSC_{phase}", 'N/A'),
             "key": f"MSC_{phase}"},
            {"label": f"Net Filter Weight {p}", "unit":  units.get(f"PMsample_mass_{phase}", 'N/A'),
             "key": f"PMsample_mass_{phase}"},
            {"label": f"Balance Cal Check {p}", "unit": units.get(f"Balance_cal_check_{phase}", "N/A"),
             "key": f"Balance_cal_check_{phase}"},
            {"label": f"Filter Mass Threshold {p}", "unit": units.get(f"filter_loading_threshhold_{phase}", "N/A"),
             "key": f"filter_loading_threshhold_{phase}"},
            {"label": f"Tare Sets Till Convergence {p}", "unit": units.get(f"Tare_sets_{phase}", "N/A"),
             "key": f"Tare_sets_{phase}"},
            {"label": f"Gross Sets Till Convergence {p}", "unit": units.get(f"Gross_sets_{phase}", "N/A"),
             "key": f"Tare_sets_{phase}"},
            {"label": f"Gravimetric A Flow Rate Check {p}",
             "unit": units.get(f"Gravimetric_A_Flow_Check_{phase}", "N/A"),
             "key": f"Gravimetric_A_Flow_Check_{phase}"},
            {"label": f"Gravimetric B Flow Rate Check {p}",
             "unit": units.get(f"Gravimetric_B_Flow_Check_{phase}", "N/A"),
             "key": f"Gravimetric_B_Flow_Check_{phase}"},
            {"label": f"Desicator Temperature {p}", "unit": units.get(f"Dessicator_temp_{phase}", "N/A"),
             "key": f"Dessicator_temp_{phase}"},
            {"label": f"Desicator Relative Humidity {p}", "unit": units.get(f"Dessicator_RH_{phase}", "N/A"),
             "key": f"Dessicator_RH_{phase}"}
        ]

        for info in PM_rows:
            ws[f'A{row_idx}'] = info["label"]
            ws[f'B{row_idx}'] = info["unit"]

            # Get values from data if available
            for test_idx in range(num_tests):
                col_letter = get_column_letter(test_idx + 3)
                cell_key = f"{info['key']}"

                if cell_key in data_values and "values" in data_values[cell_key]:
                    try:
                        value = round(float(data_values[cell_key]["values"][test_idx]), 3)
                        ws[f'{col_letter}{row_idx}'] = value
                    except IndexError:
                        ws[f'{col_letter}{row_idx}'] = ""
                    except (TypeError, ValueError):
                        value = data_values[cell_key]["values"][test_idx]
                        ws[f'{col_letter}{row_idx}'] = value
                else:
                    ws[f'{col_letter}{row_idx}'] = ""

            # Apply borders to all cells in row
            for col_idx in range(1, num_tests + 3):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1
    # === DILUTION TUNNEL QUALITY CONTROL ===
    #  header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "Dilution Tunnel Quality Control"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    dil_rows = [
        {"label": "Hood Total Capture", "unit": units.get("Hood_Total_Capture_Check", 'N/A'),
         "key": "Hood_Total_Capture_Check"},
        {"label": "Flow Grid Calibration Factor", "unit": units.get("flowgrid_cal_factor", 'N/A'),
         "key": "flowgrid_cal_factor"},
        {"label": "Negative Pressure Leak Rate", "unit": units.get("Negative_Pressure_Sensor_Leak_Rate", 'N/A'),
         "key": "Negative_Pressure_Sensor_Leak_Rate"},
        {"label": "Negative Pressure Leak Check", "unit": units.get("Negative_Pressure_Sensor_Leak_Check", 'N/A'),
         "key": "Negative_Pressure_Sensor_Leak_Check"},
        {"label": "Positive Pressure Leak Rate", "unit": units.get("Positive_Pressure_Sensor_Leak_Rate", 'N/A'),
         "key": "Positive_Pressure_Sensor_Leak_Rate"},
        {"label": "Positive Pressure Leak Check", "unit": units.get("Positive_Pressure_Sensor_Leak_Check", 'N/A'),
         "key": "Positive_Pressure_Sensor_Leak_Check"},
        {"label": "Static Pressure of Dilution Tunnel", "unit": units.get(f"static_pressure_dil_tunnel", "N/A"),
         "key": "static_pressure_dil_tunnel"}
    ]

    for info in dil_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    for phase in phases:
        #  header
        ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
        if phase == 'hp':
            p = "High Power"
            ws[f'A{row_idx}'] = "High Power"
        elif phase == 'mp':
            p = "Medium Power"
            ws[f'A{row_idx}'] = "Medium Power"
        elif phase == 'lp':
            p = "Low Power"
            ws[f'A{row_idx}'] = "Low Power"
        ws[f'A{row_idx}'].font = Font(bold=True)
        ws[f'A{row_idx}'].fill = subheader_fill
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row_idx}'].border = thin_border
        for col_idx in range(2, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
        row_idx += 1

        dil_rows = [
            {"label": f"Dilution Tunnel Average Flow Rate {p}",
             "unit": units.get(f"dilution_tunnel_flow_{phase}", "N/A"),
             "key": f"dilution_tunnel_flow_{phase}"},
            {"label": f"Dilution Tunnel Standard Deviation {p}",
             "unit": units.get(f"dilution_tunnel_flow_standard_dev_{phase}", "N/A"),
             "key": f"dilution_tunnel_flow_standard_dev_{phase}"},
            {"label": f"Dilution Tunnel Flow Check", "unit": units.get(f"flow_rate_threshold_{phase}", "N/A"),
             "key": f"flow_rate_threshold_{phase}"}
        ]

        for info in dil_rows:
            ws[f'A{row_idx}'] = info["label"]
            ws[f'B{row_idx}'] = info["unit"]

            # Get values from data if available
            for test_idx in range(num_tests):
                col_letter = get_column_letter(test_idx + 3)
                cell_key = f"{info['key']}"

                if cell_key in data_values and "values" in data_values[cell_key]:
                    try:
                        try:
                            value = round(float(data_values[cell_key]["values"][test_idx]), 3)
                        except ValueError:
                            value = data_values[cell_key]["values"][test_idx]
                        ws[f'{col_letter}{row_idx}'] = value
                    except IndexError:
                        ws[f'{col_letter}{row_idx}'] = ""
                else:
                    ws[f'{col_letter}{row_idx}'] = ""

            # Apply borders to all cells in row
            for col_idx in range(1, num_tests + 3):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

    # === GAS SENSOR QUALITY CONTROL SECTION ===

    # Gas Sensor header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "Gas Sensor Quality Control"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    # Gas sensor rows
    gas_sensor_rows = [
        {"label": "Gas Sensor Leak Rate", "unit": units.get("gas_sensor_leak_rate", 'N/A'),
         "key": "gas_sensor_leak_rate"},
        {"label": "Gas Sensor Leak Check", "unit": "Pass/Fail", "key": "gas_leak_check"}
    ]

    for info in gas_sensor_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    # === CO SECTION ===

    # CO header
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "CO"
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].fill = subheader_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].border = thin_border
    for col_idx in range(2, num_tests + 3):
        cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
        cell.border = thin_border
    row_idx += 1

    # CO rows
    co_rows = [
        {"label": "Zero Bias", "unit": units.get("zero_bias_co", 'N/A'), "key": "zero_bias_co"},
        {"label": "Span Bias", "unit": units.get("span_bias_co", 'N/A'), "key": "span_bias_co"},
        {"label": "Zero Drift", "unit": units.get("zero_drift_co", 'N/A'), "key": "zero_drift_co"},
        {"label": "Span Drift", "unit": units.get("span_drift_co", 'N/A'), "key": "span_drift_co"},
        {"label": "Zero Bias QC", "unit": "Pass/Fail", "key": "zero_bias_check_co"},
        {"label": "Span Bias QC", "unit": "Pass/Fail", "key": "span_bias_check_co"},
        {"label": "Zero Drift QC", "unit": "Pass/Fail", "key": "zero_drift_check_co"},
        {"label": "Span Drift QC", "unit": "Pass/Fail", "key": "span_drift_check_co"}
    ]

    for info in co_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    # === CO2 SECTION ===

    # CO2 header
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "CO2"
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].fill = subheader_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].border = thin_border
    for col_idx in range(2, num_tests + 3):
        cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
        cell.border = thin_border
    row_idx += 1

    # CO2 rows
    co2_rows = [
        {"label": "Zero Bias", "unit": units.get("zero_bias_co2", 'N/A'), "key": "zero_bias_co2"},
        {"label": "Span Bias", "unit": units.get("span_bias_co2", 'N/A'), "key": "span_bias_co2"},
        {"label": "Zero Drift", "unit": units.get("zero_drift_co2", 'N/A'), "key": "zero_drift_co2"},
        {"label": "Span Drift", "unit": units.get("span_drift_co2", 'N/A'), "key": "span_drift_co2"},
        {"label": "Zero Bias QC", "unit": "Pass/Fail", "key": "zero_bias_check_co2"},
        {"label": "Span Bias QC", "unit": "Pass/Fail", "key": "span_bias_check_co2"},
        {"label": "Zero Drift QC", "unit": "Pass/Fail", "key": "zero_drift_check_co2"},
        {"label": "Span Drift QC", "unit": "Pass/Fail", "key": "span_drift_check_co2"}
    ]

    for info in co2_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    # === ENVIRONMENTAL QUALITY CONTROL ===
    #  header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "Environmental Quality Control"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    env_rows = [
        {"label": "Initial Wind Speed", "unit": units.get("initial_wind_velocity", 'N/A'),
         "key": "initial_wind_velocity"},
        {"label": "Final Wind Speed", "unit": units.get("final_wind_velocity", 'N/A'),
         "key": "final_wind_velocity"},
        {"label": "Wind Speed Check", "unit": units.get("wind_speed_check", 'N/A'),
         "key": "wind_speed_check"},
        {"label": "Initial Room Temperature", "unit": units.get("initial_air_temp", 'N/A'),
         "key": "initial_air_temp"},
        {"label": "Final Room Temperature", "unit": units.get("final_air_temp", 'N/A'),
         "key": "final_air_temp"},
        {"label": "Temperature Check", "unit": units.get("temperature_check", 'N/A'),
         "key": "temperature_check"}
    ]

    for info in env_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    #########################################################################
    # Create bkg and carbon balance tab
    row_idx = 0
    ws = wb.create_sheet(title="Bkg Em and Carbon Balance")

    # Determine number of tests from data
    num_tests = 1  # Default
    for key in data_values:
        if "values" in data_values[key]:
            test_count = len(data_values[key]["values"])
            if test_count > 0:
                num_tests = test_count
                break

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    for col_idx in range(3, num_tests + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    row_idx += 1

    # === Background emissions ===
    #  header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "Background Emissions"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    bkg_rows = [
        {"label": "CO Concentration Before Test", "unit": units.get("CO_prebkg_conc", 'N/A'),
         "key": "CO_prebkg_conc"},
        {"label": "CO2 Concentration Before Test", "unit": units.get("CO2_prebkg_conc", 'N/A'),
         "key": "CO2_prebkg_conc"},
        {"label": "PM Concentration Before Test", "unit": units.get("PM_prebkg_conc", 'N/A'),
         "key": "PM_prebkg_conc"},
        {"label": "CO Concentration After Test", "unit": units.get("CO_postbkg_conc", 'N/A'),
         "key": "CO_postbkg_conc"},
        {"label": "CO2 Concentration After Test", "unit": units.get("CO2_postbkg_conc", 'N/A'),
         "key": "CO2_postbkg_conc"},
        {"label": "PM Concentration After Test", "unit": units.get("PM_postbkg_conc", 'N/A'),
         "key": "PM_postbkg_conc"}
    ]

    for info in bkg_rows:
        ws[f'A{row_idx}'] = info["label"]
        ws[f'B{row_idx}'] = info["unit"]

        # Get values from data if available
        for test_idx in range(num_tests):
            col_letter = get_column_letter(test_idx + 3)
            cell_key = f"{info['key']}"

            if cell_key in data_values and "values" in data_values[cell_key]:
                try:
                    try:
                        value = round(float(data_values[cell_key]["values"][test_idx]), 3)
                    except ValueError:
                        value = data_values[cell_key]["values"][test_idx]
                    ws[f'{col_letter}{row_idx}'] = value
                except IndexError:
                    ws[f'{col_letter}{row_idx}'] = ""
            else:
                ws[f'{col_letter}{row_idx}'] = ""

        # Apply borders to all cells in row
        for col_idx in range(1, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        row_idx += 1

    #=== CARBON BALANCE ===
    #  header row
    ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
    ws[f'A{row_idx}'] = "Carbon Balance"
    ws[f'A{row_idx}'].fill = header_fill
    ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row_idx}'].font = Font(bold=True)
    ws[f'A{row_idx}'].border = thin_border
    ws[f'{get_column_letter(num_tests + 2)}{row_idx}'].border = thin_border
    row_idx += 1

    for phase in phases:
        #  header
        ws.merge_cells(f'A{row_idx}:{get_column_letter(num_tests + 2)}{row_idx}')
        if phase == 'hp':
            p = "High Power"
            ws[f'A{row_idx}'] = "High Power"
        elif phase == 'mp':
            p = "Medium Power"
            ws[f'A{row_idx}'] = "Medium Power"
        elif phase == 'lp':
            p = "Low Power"
            ws[f'A{row_idx}'] = "Low Power"
        ws[f'A{row_idx}'].font = Font(bold=True)
        ws[f'A{row_idx}'].fill = subheader_fill
        ws[f'A{row_idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{row_idx}'].border = thin_border
        for col_idx in range(2, num_tests + 3):
            cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
            cell.border = thin_border
        row_idx += 1

        carb_rows = [
            {"label": f"Carbon In {p}", "unit": units.get(f"carbon_in_{phase}", "N/A"), "key": f"carbon_in_{phase}"},
            {"label": f"Carbon Out {p}", "unit": units.get(f"carbon_out_{phase}", "N/A"), "key": f"carbon_out_{phase}"},
            {"label": f"Carbon In to Carbon Out", "unit": units.get(f"C_Out_In_{phase}", "N/A"),
             "key": f"C_Out_In_{phase}"}
        ]

        for info in carb_rows:
            ws[f'A{row_idx}'] = info["label"]
            ws[f'B{row_idx}'] = info["unit"]

            # Get values from data if available
            for test_idx in range(num_tests):
                col_letter = get_column_letter(test_idx + 3)
                cell_key = f"{info['key']}"

                if cell_key in data_values and "values" in data_values[cell_key]:
                    try:
                        try:
                            value = round(float(data_values[cell_key]["values"][test_idx]), 3)
                        except ValueError:
                            value = data_values[cell_key]["values"][test_idx]
                        ws[f'{col_letter}{row_idx}'] = value
                    except IndexError:
                        ws[f'{col_letter}{row_idx}'] = ""
                else:
                    ws[f'{col_letter}{row_idx}'] = ""

            # Apply borders to all cells in row
            for col_idx in range(1, num_tests + 3):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

    # Save the workbook
    wb.save(outputpath)
    line = f"Created ISO Excel Report: {outputpath}"
    print(line)
    logs.append(line)

    io.write_logfile(logpath, logs)

